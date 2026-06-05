"""Generate SPARK_BOM.xlsx — console bill of materials: prices, opinion, 2026 flagship comparison, buy links."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

NCOLS = 10  # A..J

wb = Workbook()
ws = wb.active
ws.title = "BOM SPARK"

ORANGE = "F08C29"; DARK = "2B2B33"; LIGHT = "FCEFE0"; LINKBLUE = "1A5FB4"

title_font = Font(size=16, bold=True, color="FFFFFF")
hdr_font = Font(size=11, bold=True, color="FFFFFF")
cell_font = Font(size=11)
bold = Font(size=11, bold=True)
link_font = Font(size=10, color=LINKBLUE, underline="single")
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def last_col():
    return get_column_letter(NCOLS)

# Title
ws.merge_cells("A1:%s1" % last_col())
ws["A1"] = "SPARK — Bill of Materials (BOM)  ·  estimations à valider via les liens"
ws["A1"].font = title_font
ws["A1"].fill = PatternFill("solid", fgColor=ORANGE)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# Note line
ws.merge_cells("A2:%s2" % last_col())
ws["A2"] = ("Prix indicatifs en €. 'Retail' = à l'unité en magasin. 'Pro/volume' = estimation fournisseur en gros (>=1000 u., tray/B2B), toujours plus bas. "
            "Liens majoritairement US (USD) -> convertir/valider. 'Phare 2026' = le meilleur équivalent du marché, pour situer le choix.")
ws["A2"].font = Font(italic=True, size=9, color="555555")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 40

headers = ["Composant", "Pièce recommandée / détail", "Mon avis (expert)",
           "Prix retail u. (€)", "Prix pro/volume (€)",
           "Équivalent phare 2026", "Avis différence (phare vs choisi)",
           "Lien 1", "Lien 2", "Lien 3"]
ws.append(headers)
hr = 3
for c in range(1, NCOLS + 1):
    cell = ws.cell(row=hr, column=c)
    cell.font = hdr_font
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[hr].height = 34

# [comp, piece, avis, retail, pro, flagship2026, diff, l1, l2, l3]
rows = [
    ["APU AMD (Ryzen 7 8845HS, 780M)",
     "APU 8C/16T + iGPU RDNA3 780M. Pas vendu nu au détail -> proxy mini-PC, ou tray AMD en volume.",
     "Coeur de la console. 8845HS = top 1080p indé mais cher -> tue le 300€. Alternatives : Z1 non-Extreme, 7640U/8640U, R5 660M. Puce NON vendue nue -> reference achetable = mini-PC complet ~799$ (Beelink SER8, prix verifie 2026).",
     230, 160,
     "Ryzen AI 9 HX 370 (Radeon 890M RDNA3.5) / AMD Z2 Extreme",
     "890M ~25-40% > 780M, mais +cher/+chaud/+conso. Pour de l'indé 1080p, 780M = sweet spot ; le phare = overkill qui casse le budget.",
     "https://www.bee-link.com/products/beelink-ser8-8845hs",
     "https://www.newegg.com/p/pl?N=100008345+601456629",
     "https://www.gmktec.com/products/amd-ryzen-7-8845hs-mini-pc-nucbox-k8"],

    ["RAM 16 Go DDR5-5600 SO-DIMM",
     "1x16 Go (ou 2x8 dual-channel = mieux iGPU). LPDDR5 soudé en volume = moins cher/plus fin.",
     "Dual-channel obligatoire pour l'iGPU. 16 Go mini. En volume, LPDDR5 soudée > SO-DIMM.",
     42, 26,
     "32 Go LPDDR5X-7500/8000 soudée",
     "Plus de bande passante = +5-15% iGPU + multitâche. 16 Go DC-5600 suffit en 1080p indé. Phare = confort, pas indispensable.",
     "https://www.newegg.com/crucial-16gb-ddr5-5600-cas-latency-cl46-notebook-memory/p/N82E16820156314",
     "https://www.newegg.com/kingston-16gb/p/0RM-001W-006C8",
     "https://www.newegg.com/p/pl?d=ddr5+5600+sodimm"],

    ["Stockage NVMe 512 Go M.2 2280",
     "SSD PCIe Gen3/Gen4. 256 Go trop juste, 512 Go bon compromis.",
     "Gen3 suffit pour de l'indé (Gen4 inutile, +cher/+chaud). 512 Go mini vu la taille des jeux.",
     45, 30,
     "1 To NVMe Gen4 (voire Gen5)",
     "Capacité x2, débit x2. Inutile pour indé (jeux légers) ; Gen4/5 = +chaud/+cher. 512 Go Gen3 OK.",
     "https://www.amazon.com/Silicon-Power-512GB-Gen3x4-SP512GBP34A60M28/dp/B07ZGJYLNL",
     "https://www.amazon.com/Western-Digital-512GB-Gaming-Internal/dp/B0CKRX3WDH",
     "https://www.newegg.com/p/pl?d=512gb+ssd+m.2"],

    ["Carte mère custom (PCBA)",
     "PCB custom + assemblage CMS. NRE (étude) une fois, puis coût/unité en volume.",
     "Le vrai poste 'console' : concevoir le PCB autour de l'APU. NRE élevé amorti sur volume. Pas d'achat retail.",
     70, 45,
     "PCB 6-8 couches, VRM renforcé, Wi-Fi 7 intégré",
     "Marginal pour ce TDP. +couches = +coût. Un 4 couches bien conçu suffit. Phare = inutile au début.",
     "https://jlcpcb.com",
     "https://www.pcbway.com",
     "https://www.alibaba.com/showroom/pcba-assembly.html"],

    ["Refroidissement (blower + caloduc)",
     "Ventilo blower bas profil + heatsink/caloduc dimensionné au TDP (35-54W).",
     "Wall-powered -> pousse le TDP, dimensionne large + silencieux. Type laptop/console.",
     16, 9,
     "Chambre à vapeur + grand ventilo (type Steam Deck OLED)",
     "Plus silencieux/froid à TDP élevé. TDP modéré -> blower+caloduc suffit. Phare = confort acoustique surtout.",
     "https://www.amazon.com/mini-pc-cooling-fan/s?k=mini+pc+cooling+fan",
     "https://www.alibaba.com/showroom/mini-pc-fan.html",
     "https://mitxpc.com/collections/cpu-coolers"],

    ["Alimentation 19V ~120W",
     "Bloc externe (brique) 120W, ou alim interne. Externe = moins cher/plus simple.",
     "Brique externe = économique, évite la chaleur interne. 120W large pour l'APU.",
     26, 12,
     "Chargeur GaN compact 140W USB-C PD",
     "Plus petit/moderne/USB-C universel (recharge accessoires). Brique barrel 120W = -cher. USB-C PD = bel argument 2026.",
     "https://www.walmart.com/ip/19V-6-3A-120W-Laptop-AC-Adapter-Power-Supply-Charger-US-Power-Cord-for-ASUS/554524924",
     "https://www.newegg.com/p/pl?d=19v+power+adapter",
     "https://www.ebay.com/b/19V-Laptop-Power-Adapters-Chargers/31510/bn_650230"],

    ["Boîtier (coque)",
     "Plastique injecté (moule custom, NRE) ou tôle/alu petit volume.",
     "Moule injection = NRE 30-100k€ mais qq € l'unité en volume. Identité produit.",
     25, 16,
     "Unibody aluminium + finitions premium",
     "Toucher/thermique premium mais NRE + coût lourds. ABS injecté = bon pour 300€. L'identité (design) compte > le matériau.",
     "https://www.alibaba.com/showroom/plastic-enclosure.html",
     "https://www.protocase.com",
     "https://www.pcbway.com/rapid-prototyping/cnc-machining/"],

    ["Lecteur micro-SD (cartouche)",
     "Connecteur SD push-push qualité + circuit. (Modules 'Arduino' = proto seulement.)",
     "Le slot 'cartouche'. Coût pièce faible. La SÉCURITÉ cartouche est un autre sujet (voir phare).",
     4, 1.5,
     "Connecteur UHS-II + élément sécurisé (secure element) pour DRM",
     "ICI le phare COMPTE : un élément sécurisé = vraie sécurité cartouche (sinon copiable). UHS-II = chargement + rapide. Surcoût lié à ton modèle cartouche -> à trancher.",
     "https://www.amazon.com/MicroSD-Breakout-Reader-Module-Expansion/dp/B09SD7D1VK",
     "https://www.ebay.com/itm/386914952215",
     "https://www.amazon.com/pzsmocn-Micro-SD-Compatible-Raspberry-Teaching/dp/B08FB3GC34"],

    ["Manette",
     "Manette sans fil (bundle) ou vendue à part. OEM Chine en volume.",
     "Si incluse -> +coût. Option : la vendre séparément pour tenir le prix. OEM customisable.",
     28, 14,
     "Sticks Hall effect + haptique + gâchettes analogiques",
     "Hall = ANTI-DRIFT (gros argument durabilité/marketing). Haptique = ressenti premium. Vaut le coup d'envisager les sticks Hall même en milieu de gamme.",
     "https://www.alibaba.com/showroom/wholesale-game-controller.html",
     "https://www.made-in-china.com/products-search/hot-china-products/Wholesale_Game_Controller.html",
     "https://www.amazon.com/Wireless-Controllers/s?k=Wireless+Controllers"],

    ["Connecteurs / câbles / divers",
     "HDMI/DP, USB, jack, nappe, vis, antenne Wi-Fi/BT, pâte thermique.",
     "Petit matériel. Sourcer via distributeurs élec en volume.",
     14, 9,
     "Connecteurs premium, Wi-Fi 7 / BT 5.4",
     "Wi-Fi 6E suffit largement. Premium = marginal, peu visible. Pas prioritaire.",
     "https://www.lcsc.com",
     "https://www.mouser.com",
     "https://www.digikey.com"],

    ["Assemblage + test",
     "Montage, flash OS, test fonctionnel, emballage. Sous-traité (EMS) en volume.",
     "Coût main d'oeuvre/EMS. Baisse fort en volume. Inclure test + flash SPARK.",
     22, 14,
     "EMS premium + QC 100% + burn-in",
     "QC renforcé = moins de SAV/retours. Coût + mais fiabilité. Compromis selon volume/budget.",
     "https://www.alibaba.com/showroom/ems-electronic-manufacturing-service.html",
     "https://jlcpcb.com/smt-assembly",
     "https://www.pcbway.com/pcb-assembly.html"],
]

r = hr + 1
for row in rows:
    ws.cell(row=r, column=1, value=row[0]).font = bold
    ws.cell(row=r, column=2, value=row[1]).font = cell_font
    ws.cell(row=r, column=3, value=row[2]).font = cell_font
    ws.cell(row=r, column=4, value=row[3]).font = cell_font
    ws.cell(row=r, column=5, value=row[4]).font = cell_font
    ws.cell(row=r, column=4).number_format = u'#,##0 €'
    ws.cell(row=r, column=5).number_format = u'#,##0 €'
    ws.cell(row=r, column=6, value=row[5]).font = cell_font
    ws.cell(row=r, column=7, value=row[6]).font = cell_font
    for i, url in enumerate(row[7:10]):
        c = ws.cell(row=r, column=8 + i, value="Lien %d" % (i + 1))
        c.hyperlink = url
        c.font = link_font
    for c in range(1, NCOLS + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if (r - hr) % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT)
    ws.row_dimensions[r].height = 66
    r += 1

# Totals
total_row = r
ws.cell(row=total_row, column=1, value="TOTAL BOM").font = Font(bold=True, size=12)
ws.cell(row=total_row, column=3, value="Somme des composants (hors NRE moule/PCB, transport, marge, R&D, marketing)").font = Font(italic=True, size=9, color="555555")
d = ws.cell(row=total_row, column=4, value="=SUM(D4:D%d)" % (total_row - 1))
e = ws.cell(row=total_row, column=5, value="=SUM(E4:E%d)" % (total_row - 1))
for cc in (d, e):
    cc.font = Font(bold=True, size=12)
    cc.number_format = u'#,##0 €'
    cc.fill = PatternFill("solid", fgColor="FFE3C2")
for c in range(1, NCOLS + 1):
    ws.cell(row=total_row, column=c).border = border
ws.row_dimensions[total_row].height = 24

# Analysis block
ar = total_row + 2
notes = [
    "LECTURE :",
    "- Pro/volume ~ somme ci-dessus. Retail u. = si tu achetais tout à l'unité (jamais en prod).",
    "- Prix de vente realiste = BOM pro x ~1.6 a 2.0 (transport + douane + marge revendeur 30% + R&D amortie + support + marketing).",
    "- Donc BOM pro ~300-340€ -> prix retail realiste ~450-600€. Le 300€ retail = TRES dur sans volume Nintendo/Valve.",
    "",
    "OU LE 'PHARE' VAUT LE COUP (selon moi) :",
    "- Manette sticks Hall (anti-drift) = vrai argument durabilite/marketing.",
    "- Slot cartouche avec element securise = SI tu veux une vraie securite cartouche (sinon copiable).",
    "- USB-C PD pour l'alim = moderne, recharge accessoires.",
    "Le reste (APU 890M, 32 Go, 1 To Gen4, alu, vapor chamber) = surcout pour gain faible sur de l'indé 1080p.",
    "",
    "POUR VISER 300€ : APU moins cher (Z1/7640U/R5 660M, -70 a -120€), LPDDR5 soudee, manette a part. OU viser 399-449€. OU vendre a perte + abo/store.",
    "",
    "ANCRAGE REEL VERIFIE (2026) : un mini-PC 8845HS COMPLET pret a l'emploi (Beelink SER8) = ~799$ (~740€).",
    "-> Meme une machine de ce niveau toute faite coute 600-800€. Donc 300€ retail neuf = encore plus dur que l'estime.",
    "",
    "FIABILITE PRIX : colonnes = ESTIMATIONS expertes. Certains liens pointent vers une machine entiere ou une categorie, pas la piece exacte -> a confirmer par de vrais devis (APU tray AMD, RAM/SSD volume, devis PCBA JLCPCB/PCBWay, devis moule boitier).",
]
for n in notes:
    cell = ws.cell(row=ar, column=1, value=n)
    cell.font = Font(bold=n.endswith(":"), size=10, color=("B5651D" if n.endswith(":") else "333333"))
    ws.merge_cells(start_row=ar, start_column=1, end_row=ar, end_column=NCOLS)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ar += 1

# Column widths
widths = [28, 34, 40, 13, 15, 30, 40, 8, 8, 8]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A4"

# ---------------------------------------------------------------------------
# Sheet 2 : APU 680M vs 780M comparison
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("APU 680M vs 780M")
ws2.merge_cells("A1:D1")
ws2["A1"] = "SPARK — APU : Radeon 680M (base) vs 780M (Pro)  ·  prix/perf indicatifs, à valider"
ws2["A1"].font = title_font
ws2["A1"].fill = PatternFill("solid", fgColor=ORANGE)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

c2_hdr = ["Critère", "Radeon 680M (base)", "Radeon 780M (Pro)", "Différence / note"]
ws2.append(c2_hdr)
for c in range(1, 5):
    cell = ws2.cell(row=2, column=c)
    cell.font = hdr_font
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws2.row_dimensions[2].height = 28

comp = [
    ["Architecture", "RDNA2", "RDNA3", "RDNA3 = dual-issue + clocks + IA"],
    ["Compute Units", "12 CU", "12 CU", "Même nombre, mais RDNA3 plus efficace"],
    ["Horloge GPU", "~2.2-2.4 GHz", "~2.7 GHz", "780M plus haut"],
    ["APU exemples", "Ryzen 6800H / 6900HX / 7735HS", "Ryzen 7840HS / 8840HS / 8845HS", "780M = génération plus récente"],
    ["Perf GPU relative", "100 % (base)", "~+25-30 %", "Écart net sur AA/AAA, faible sur le reste"],
    ["Indé / 2D / esport 1080p", "60 fps facile", "60 fps facile", "ÉGAL (les deux suffisent)"],
    ["AA / un peu vieux 1080p", "30-45 fps medium", "45-60 fps medium", "780M plus confortable"],
    ["AAA récents 1080p", "720p low + FSR ~30", "1080p low + FSR 30-45", "780M = un cran au-dessus"],
    ["Émulation GC/Wii/PS2", "OK", "OK + marge (réso interne +)", "Les deux bien"],
    ["Upscaling FSR / AFMF", "Identique", "Identique", "PAS lié à la puissance"],
    ["TDP / conso", "~35-54 W", "~35-54 W", "Similaire (wall-powered)"],
    ["Prix plateforme (BOM, volume est.)", "~90-130 €", "~150-200 €", "Δ ~ +40-70 € pour le 780M"],
    ["Mini-PC proxy (retail)", "~300-400 € (Beelink SER6…)", "~500-800 € (Beelink SER8…)", "Machine entière, pas la puce seule"],
]
r = 3
for row in comp:
    ws2.cell(row=r, column=1, value=row[0]).font = bold
    for c in range(2, 5):
        ws2.cell(row=r, column=c, value=row[c - 1]).font = cell_font
    for c in range(1, 5):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r % 2 == 1:
            cell.fill = PatternFill("solid", fgColor=LIGHT)
    ws2.row_dimensions[r].height = 30
    r += 1

# verdict block
vr = r + 1
verdict = [
    "VERDICT :",
    "- Pour de l'INDÉ / esport / rétro / création (FORGE) : 680M = SUFFISANT → choisis-le (moins cher, ~-40-70 € BOM, plus près de 300 €).",
    "- Pour viser AA/AAA confortables : 780M (= SKU 'Pro', prix plus haut).",
    "- L'upscaling (FSR/AFMF) est IDENTIQUE sur les deux → le 680M en profite autant.",
    "- Δ prix puce ~40-70 € en volume ; les écarts mini-PC (300 vs 800 €) = machines entières, pas la puce.",
    "- Stratégie SPARK : 680M = base abordable, 780M = Pro optionnel. (Premium AAA = Strix Halo, autre gamme.)",
    "Prix = ESTIMATIONS à confirmer par devis (AMD tray / fabricant).",
]
for v in verdict:
    cell = ws2.cell(row=vr, column=1, value=v)
    cell.font = Font(bold=v.endswith(":"), size=10, color=("B5651D" if v.endswith(":") else "333333"))
    ws2.merge_cells(start_row=vr, start_column=1, end_row=vr, end_column=4)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    vr += 1

for i, w in enumerate([34, 30, 32, 40], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A3"

# ---------------------------------------------------------------------------
# Sheet 3 : full machine cost — 680M build vs 780M build
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Coût machine 680M vs 780M")
ws3.merge_cells("A1:D1")
ws3["A1"] = "SPARK — Coût machine complète : build 680M vs build 780M  ·  prix pro/volume estimés (€)"
ws3["A1"].font = title_font
ws3["A1"].fill = PatternFill("solid", fgColor=ORANGE)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

ws3.append(["Composant", "Build 680M (€)", "Build 780M (€)", "Note"])
for c in range(1, 5):
    cell = ws3.cell(row=2, column=c)
    cell.font = hdr_font
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws3.row_dimensions[2].height = 26

# [composant, prix680, prix780, note]
mc = [
    ["APU", 110, 170, "Seule vraie différence (680M vs 780M, volume)"],
    ["RAM 16 Go DDR5 dual-channel", 26, 26, ""],
    ["Stockage NVMe 512 Go", 30, 30, ""],
    ["Carte mère custom (PCBA)", 45, 45, "hors NRE étude"],
    ["Refroidissement", 9, 9, ""],
    ["Alimentation", 12, 12, ""],
    ["Boîtier", 16, 16, "hors NRE moule"],
    ["Lecteur micro-SD", 1.5, 1.5, ""],
    ["Manette", 14, 14, "option : vendre à part"],
    ["Connecteurs / divers", 9, 9, ""],
    ["Assemblage + test", 14, 14, ""],
]
r = 3
for row in mc:
    ws3.cell(row=r, column=1, value=row[0]).font = bold
    ws3.cell(row=r, column=2, value=row[1]).font = cell_font
    ws3.cell(row=r, column=3, value=row[2]).font = cell_font
    ws3.cell(row=r, column=4, value=row[3]).font = Font(size=9, italic=True, color="666666")
    ws3.cell(row=r, column=2).number_format = u'#,##0 €'
    ws3.cell(row=r, column=3).number_format = u'#,##0 €'
    for c in range(1, 5):
        cell = ws3.cell(row=r, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r % 2 == 1:
            cell.fill = PatternFill("solid", fgColor=LIGHT)
    r += 1

# TOTAL BOM
tr = r
ws3.cell(row=tr, column=1, value="TOTAL BOM (pièces)").font = Font(bold=True, size=12)
b680 = ws3.cell(row=tr, column=2, value="=SUM(B3:B%d)" % (tr - 1))
b780 = ws3.cell(row=tr, column=3, value="=SUM(C3:C%d)" % (tr - 1))
for cc in (b680, b780):
    cc.font = Font(bold=True, size=12)
    cc.number_format = u'#,##0 €'
    cc.fill = PatternFill("solid", fgColor="FFE3C2")
ws3.cell(row=tr, column=4, value="hors transport/marge/R&D/marketing").font = Font(size=9, italic=True, color="666666")
for c in range(1, 5):
    ws3.cell(row=tr, column=c).border = border

# Retail estimates
r2 = tr + 1
ws3.cell(row=r2, column=1, value="Prix retail estimé ×1.5 (direct, marge fine)").font = bold
rb1 = ws3.cell(row=r2, column=2, value="=B%d*1.5" % tr)
rb2 = ws3.cell(row=r2, column=3, value="=C%d*1.5" % tr)
r3 = tr + 2
ws3.cell(row=r3, column=1, value="Prix retail estimé ×1.8 (marge + distrib)").font = bold
rc1 = ws3.cell(row=r3, column=2, value="=B%d*1.8" % tr)
rc2 = ws3.cell(row=r3, column=3, value="=C%d*1.8" % tr)
for cc in (rb1, rb2, rc1, rc2):
    cc.number_format = u'#,##0 €'
    cc.font = bold
for rr in (r2, r3):
    for c in range(1, 5):
        ws3.cell(row=rr, column=c).border = border

# notes
nr = r3 + 2
mnotes = [
    "LECTURE :",
    "- Seule différence = l'APU (680M ~110€ vs 780M ~170€) -> Δ BOM ~ +60€ pour le 780M.",
    "- TOTAL BOM ~ 287€ (680M) vs ~347€ (780M). Retail réaliste = BOM ×1.5 à ×1.8.",
    "- 680M : retail ~430-516€. 780M : retail ~520-625€.",
    "- => Le 300€ retail N'EST PAS atteint, même en 680M (BOM déjà ~287€). Pour 300€ : APU encore plus faible (Vega/660M), couper RAM/manette, ou vendre à perte + abo.",
    "- Manette vendue à part = -14€ BOM. APU plus faible = la plus grosse économie.",
    "Prix = ESTIMATIONS (à confirmer par devis AMD tray / fabricant / PCBA / moule).",
]
for n in mnotes:
    cell = ws3.cell(row=nr, column=1, value=n)
    cell.font = Font(bold=n.endswith(":"), size=10, color=("B5651D" if n.endswith(":") else "333333"))
    ws3.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=4)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    nr += 1

for i, w in enumerate([34, 16, 16, 38], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A3"

out = os.path.join(os.path.dirname(__file__), "SPARK_BOM.xlsx")
wb.save(out)
print("saved", out)
